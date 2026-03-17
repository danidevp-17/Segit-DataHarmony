"""
Preview de archivos para Data Quality.
Retorna JSON compatible con DocViewerPanel (type, content, sheets, html, etc.).
"""
import io
from typing import Any

MAX_TEXT_BYTES = 2 * 1024 * 1024

TEXT_MIMES = {
    "text/plain", "text/csv", "text/html", "application/json",
    "application/sql", "application/x-sh", "text/x-python", "text/x-shellscript",
}


def _is_text_mime(mime: str) -> bool:
    return mime in TEXT_MIMES or mime.startswith("text/")


def _is_image_mime(mime: str) -> bool:
    return mime.startswith("image/")


def _image_exts() -> set[str]:
    return {"png", "jpg", "jpeg", "gif", "webp", "bmp", "svg"}


def _pptx_mimes() -> set[str]:
    return {
        "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        "application/vnd.openxmlformats-officedocument.presentationml.template",
        "application/vnd.ms-powerpoint",
    }


def _pptx_exts() -> set[str]:
    return {"pptx", "potx", "ppt", "pot"}


def _xlsx_mimes() -> set[str]:
    return {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    }


def _xlsx_exts() -> set[str]:
    return {"xlsx", "xls", "csv"}


def _docx_mimes() -> set[str]:
    return {"application/vnd.openxmlformats-officedocument.wordprocessingml.document"}


def get_preview(file_id: str, buffer: bytes, mime_type: str, file_name: str) -> dict[str, Any]:
    """Genera el JSON de preview."""
    ext = (file_name.split(".")[-1] if file_name else "").lower()

    # Images
    if _is_image_mime(mime_type) or ext in _image_exts():
        return {"type": "image"}

    # PDF
    if mime_type == "application/pdf" or ext == "pdf":
        return {"type": "pdf"}

    # PPTX
    if mime_type in _pptx_mimes() or ext in _pptx_exts():
        return {
            "type": "pptx_legacy",
            "fileName": file_name,
            "info": f"La vista previa de archivos PowerPoint ({ext.upper()}) no está disponible. Descarga el archivo para abrirlo.",
        }

    # XLSX / CSV
    if mime_type in _xlsx_mimes() or ext in _xlsx_exts():
        try:
            import openpyxl
            if ext == "csv" or mime_type == "text/csv":
                import csv
                text = buffer[:MAX_TEXT_BYTES].decode("utf-8", errors="replace")
                reader = csv.reader(io.StringIO(text))
                rows = list(reader)
                return {"type": "xlsx", "sheetNames": ["Sheet1"], "sheets": {"Sheet1": rows}}
            wb = openpyxl.load_workbook(io.BytesIO(buffer), read_only=True)
            sheets = {}
            for name in wb.sheetnames:
                ws = wb[name]
                rows = [[cell.value for cell in row] for row in ws.iter_rows(max_row=500)]
                sheets[name] = rows
            wb.close()
            return {"type": "xlsx", "sheetNames": wb.sheetnames, "sheets": sheets}
        except Exception:
            return {"type": "unsupported", "mimeType": mime_type, "fileName": file_name}

    # DOCX
    if mime_type in _docx_mimes() or ext == "docx":
        try:
            import mammoth
            result = mammoth.convert_to_html(io.BytesIO(buffer))
            return {"type": "docx", "html": result.value}
        except Exception:
            return {"type": "unsupported", "mimeType": mime_type, "fileName": file_name}

    # DLIS
    if ext == "dlis":
        return {
            "type": "pptx_legacy",
            "fileName": file_name,
            "info": "DLIS es un formato binario. No es posible previsualizarlo.",
        }

    # LAS / LIS
    if ext in ("las", "lis"):
        raw = buffer[:MAX_TEXT_BYTES]
        text = raw.decode("latin1", errors="replace")
        if len(buffer) > MAX_TEXT_BYTES:
            text += "\n\n[...archivo truncado...]"
        return {"type": "text", "content": text, "ext": ext}

    # Text / code
    if _is_text_mime(mime_type) or ext in {"txt", "md", "py", "sh", "sql", "js", "ts", "json"}:
        raw = buffer[:MAX_TEXT_BYTES]
        text = raw.decode("utf-8", errors="replace")
        if len(buffer) > MAX_TEXT_BYTES:
            text += "\n\n[...archivo truncado...]"
        return {"type": "text", "content": text, "ext": ext}

    return {"type": "unsupported", "mimeType": mime_type, "fileName": file_name}
